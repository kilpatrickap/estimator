import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Protection

class SubcontractorIO:
    """Handles exporting and importing Excel RFQs securely to ensure data integrity."""

    @staticmethod
    def export_rfq(db_path, package_name, output_path, items_data):
        """
        Exports a locked Excel file for the subcontractor.
        Uses items_data (list of dicts) representing the current package view.
        """
        if not items_data:
            raise ValueError("No items found for the selected package.")

        # 1. Prepare Data for Pandas DataFrame
        data = []
        for row in items_data:
            data.append({
                "Internal_ID": row.get('rowid', ''), # The critical hidden key
                "Ref": row.get('ref', ''),
                "Description": row.get('desc', ''),
                "Qty": row.get('qty', ''),
                "Unit": row.get('unit', ''),
                "Rate": "", # Left blank for the subbee
                "Amount": "", # Can be formula or blank
                "_is_target": row.get('is_target_pkg', False) # internal flag for pandas iteration
            })

        df = pd.DataFrame(data)

        # 2. Save directly via pandas to openpyxl writer to manipulate styles
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # We don't want to export our internal flag to the final Excel file
            export_df = df.drop(columns=['_is_target'])
            export_df.to_excel(writer, index=False, sheet_name="RFQ")
            
            workbook = writer.book
            worksheet = writer.sheets["RFQ"]

            # 3. Protect Sheet to prevent tampering, but allow selecting/formatting
            worksheet.protection.sheet = True
            worksheet.protection.formatCells = False
            worksheet.protection.formatColumns = False

            # Colors and Styles
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            readonly_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") # subtle gray
            heading_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") # light green (like PBOQ)
            input_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # Light yellow

            # Format Headers
            for col in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # 4. Hide Internal ID column (Column A / 1)
            worksheet.column_dimensions['A'].hidden = True

            # --- PRE-PASS: Identify Relevant Headings ---
            relevant_headings = set()
            
            # Phase 1: Forward Scan with "Bucket fix"
            current_heading_block = []
            for row_idx in range(2, len(df) + 2):
                df_row = df.iloc[row_idx - 2]
                qty = df_row['Qty']
                is_target = df_row['_is_target']
                
                is_heading = not str(qty).strip() or str(qty).lower() == 'nan'
                
                if is_heading:
                    # 1. Look-ahead: Add heading to bucket
                    current_heading_block.append(row_idx)
                else:
                    if is_target:
                        # Target hit. All bucket headings are relevant to this target.
                        for h_idx in current_heading_block:
                            relevant_headings.add(h_idx)
                        current_heading_block = [] # Reset bucket
                    else:
                        # 2. My Fix: Non-target item hit. Must dump the bucket because these headings belong to this non-target item!
                        current_heading_block = []
                        
            # Phase 2: Look-Behind Validation
            for row_idx in range(2, len(df) + 2):
                df_row = df.iloc[row_idx - 2]
                is_target = df_row['_is_target']
                
                if is_target:
                    # 3. Look-behind: Walk backwards up the BOQ starting from the row immediately above this target
                    j = row_idx - 1
                    while j >= 2:
                        prev_row = df.iloc[j - 2]
                        prev_qty = prev_row['Qty']
                        
                        is_prev_heading = not str(prev_qty).strip() or str(prev_qty).lower() == 'nan'
                        
                        if is_prev_heading:
                            # Found a heading! Validate it into the relevant set and keep climbing upwards
                            relevant_headings.add(j)
                            j -= 1
                        else:
                            # We hit a row with a quantity (a priced item). We have exited the heading block. Stop looking behind.
                            break                    
            # --- MAIN EXPORT PASS ---
            for row_idx in range(2, len(df) + 2):
                df_row = df.iloc[row_idx - 2]
                qty = df_row['Qty']
                is_target = df_row['_is_target']
                
                is_heading = not str(qty).strip() or str(qty).lower() == 'nan'
                
                # Rule 1: Structural Rows (Headings, Notes, Blanks) -> Qty is empty
                if is_heading:
                    if row_idx in relevant_headings:
                        # It's a RELEVANT heading/note. Make it visible, give it a slight heading color, lock everything
                        for col_idx in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.protection = Protection(locked=True)
                            cell.fill = heading_fill
                            
                            # Default top alignment, plus wrap for description and bolding
                            if col_idx == 3: 
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                            else:
                                cell.alignment = Alignment(vertical="top")
                    else:
                        # NOISE Heading: Hide it completely
                        worksheet.row_dimensions[row_idx].hidden = True
                        for col_idx in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.protection = Protection(locked=True)
                        
                # Rule 2: Target Package Items -> Has Qty AND is Target
                elif is_target:
                    for col_idx in range(1, len(export_df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        # Default top alignment, plus wrap for description
                        if col_idx == 3:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        else:
                            cell.alignment = Alignment(vertical="top")
                            
                        if col_idx == 6: # Column F: 'Rate' -> Unlocked for input
                            cell.protection = Protection(locked=False)
                            cell.fill = input_fill
                        elif col_idx == 7: # Column G: 'Amount' -> Formula
                            cell.protection = Protection(locked=True)
                            cell.fill = readonly_fill
                            # qty * rate
                            cell.value = f"=IF(ISNUMBER(D{row_idx})*ISNUMBER(F{row_idx}), D{row_idx}*F{row_idx}, \"\")"
                            cell.number_format = '#,##0.00'
                        else:
                            cell.protection = Protection(locked=True)
                            cell.fill = readonly_fill
                            
                # Rule 3: Other Trades' Items -> Has Qty BUT is NOT Target
                else:
                    # Hide the entire row
                    worksheet.row_dimensions[row_idx].hidden = True
                    for col_idx in range(1, len(export_df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.protection = Protection(locked=True)

            # 6. Adjust Widths for better viewing
            worksheet.column_dimensions['B'].width = 10 # Ref
            worksheet.column_dimensions['C'].width = 75 # Description (wider for headings)
            worksheet.column_dimensions['D'].width = 12 # Qty
            worksheet.column_dimensions['E'].width = 10 # Unit
            worksheet.column_dimensions['F'].width = 18 # Rate
            worksheet.column_dimensions['G'].width = 20 # Amount

    @staticmethod
    def import_rfq(db_path, package_name, excel_path, subcontractor_name):
        """
        Reads the filled Excel RFQ and imports rates into the database.
        Returns the number of successfully imported rates.
        """
        try:
            # Read excel using pandas
            df = pd.read_excel(excel_path, sheet_name="RFQ")
        except Exception as e:
            raise ValueError(f"Failed to read Excel file. Ensure it is a valid RFQ template.\n{e}")

        required_cols = ["Internal_ID", "Rate"]
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"Invalid Template: Missing critical column '{col}'. Please only import files generated by this system.")

        updates = 0
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        try:
            # Begin transaction
            cursor.execute("BEGIN TRANSACTION")
            
            for index, row in df.iterrows():
                rowid = row['Internal_ID']
                rate_val = row['Rate']
                
                # Skip invalid rowids or empty rates
                if pd.isna(rowid) or pd.isna(rate_val):
                    continue
                    
                try:
                    # Clean and parse rate
                    if isinstance(rate_val, str):
                        rate_clean = float(rate_val.replace(',', '').replace(' ', ''))
                    else:
                        rate_clean = float(rate_val)
                        
                    if rate_clean <= 0:
                        continue
                        
                    # 1. Check if a quote already exists for this exact item and subcontractor
                    cursor.execute("SELECT id FROM subcontractor_quotes WHERE package_name=? AND row_idx=? AND subcontractor_name=?", 
                                   (package_name, int(rowid), subcontractor_name))
                    res = cursor.fetchone()
                    
                    if res:
                        # Update existing
                        cursor.execute("UPDATE subcontractor_quotes SET rate=? WHERE id=?", (rate_clean, res[0]))
                    else:
                        # Insert new
                        cursor.execute("INSERT INTO subcontractor_quotes (package_name, row_idx, subcontractor_name, rate) VALUES (?, ?, ?, ?)", 
                                       (package_name, int(rowid), subcontractor_name, rate_clean))
                    
                    updates += 1
                except (ValueError, TypeError):
                    continue # Ignore parsing errors on a per-cell basis

            conn.commit()
        except sqlite3.Error as e:
            conn.rollback()
            raise ValueError(f"Database error during import: {e}")
        finally:
            conn.close()
            
        return updates
