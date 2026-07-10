"""
Resources Schedule (RS) Excel Exporter — Exports the aggregated Resources
Schedule to a styled .xlsx workbook with separate worksheets for each
resource type and a summary sheet.

Uses openpyxl (same dependency as the existing PBOQ exporter).
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class RSExcelExporter:
    """Exports an RSResult to a styled Excel workbook."""

    # Header styling
    HEADER_FILL = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid") if HAS_OPENPYXL else None
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10) if HAS_OPENPYXL else None

    # Resource type fills
    MATERIAL_FILL = PatternFill(start_color="f3e5f5", end_color="f3e5f5", fill_type="solid") if HAS_OPENPYXL else None
    LABOR_FILL = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid") if HAS_OPENPYXL else None
    EQUIPMENT_FILL = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid") if HAS_OPENPYXL else None
    PLANT_FILL = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid") if HAS_OPENPYXL else None
    SUMMARY_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid") if HAS_OPENPYXL else None

    # Alt row fill
    ALT_FILL = PatternFill(start_color="F5F7F9", end_color="F5F7F9", fill_type="solid") if HAS_OPENPYXL else None

    # Summary header styling
    SUMMARY_HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid") if HAS_OPENPYXL else None
    SUMMARY_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10) if HAS_OPENPYXL else None

    THIN_BORDER = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD")
    ) if HAS_OPENPYXL else None

    TOTAL_FONT = Font(name="Arial", bold=True, size=10) if HAS_OPENPYXL else None
    TOTAL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") if HAS_OPENPYXL else None

    # Number formats
    QTY_FORMAT = '#,##0.00'
    COST_FORMAT = '#,##0.00'

    def __init__(self, rs_result):
        """
        Args:
            rs_result: RSResult from RSGenerator.generate()
        """
        self.result = rs_result

    def export(self, output_path):
        """
        Exports the RS to an Excel workbook.

        Args:
            output_path: Destination .xlsx file path.

        Returns:
            (True, message) on success, (False, error_message) on failure.
        """
        if not HAS_OPENPYXL:
            return False, "openpyxl is not installed. Run: pip install openpyxl"

        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            # Write resource sheets
            self._write_resource_sheet(wb, "Materials", self.result.materials,
                                        self.MATERIAL_FILL, "Material Name")
            self._write_resource_sheet(wb, "Labour", self.result.labor,
                                        self.LABOR_FILL, "Trade")
            self._write_resource_sheet(wb, "Equipment", self.result.equipment,
                                        self.EQUIPMENT_FILL, "Equipment Name")
            self._write_resource_sheet(wb, "Plant", self.result.plant,
                                        self.PLANT_FILL, "Plant Name")

            # Write summary sheet
            self._write_summary_sheet(wb)

            # Write skipped items sheet (if any)
            if self.result.skipped_rows:
                self._write_skipped_sheet(wb)

            wb.save(output_path)
            sheet_count = len(wb.sheetnames)
            return True, f"Resources Schedule exported ({sheet_count} sheets) to:\n{output_path}"

        except Exception as e:
            return False, f"Export failed: {e}"

    # ── Resource Sheet ────────────────────────────────────────────────────

    def _write_resource_sheet(self, wb, sheet_name, entries, type_fill, name_header):
        """Writes a single resource type worksheet."""
        ws = wb.create_sheet(title=sheet_name)

        headers = ["#", name_header, "Unit", "Total Qty",
                    "Currency", "Avg. Unit Rate", "Total Cost", "Used In (Rate Codes)"]
        col_widths = [6, 40, 10, 16, 12, 16, 18, 40]

        # Write header row
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze top row
        ws.freeze_panes = "A2"

        # Write data rows
        for row_idx, entry in enumerate(entries, 2):
            is_alt = (row_idx % 2 == 0)
            row_fill = self.ALT_FILL if is_alt else None

            # #
            cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            cell.alignment = Alignment(horizontal="center")
            if row_fill:
                cell.fill = row_fill
            cell.border = self.THIN_BORDER

            # Name
            cell = ws.cell(row=row_idx, column=2, value=entry.name)
            cell.fill = type_fill
            cell.border = self.THIN_BORDER

            # Unit
            cell = ws.cell(row=row_idx, column=3, value=entry.unit)
            cell.alignment = Alignment(horizontal="center")
            if row_fill:
                cell.fill = row_fill
            cell.border = self.THIN_BORDER

            # Total Qty
            cell = ws.cell(row=row_idx, column=4, value=round(entry.total_qty, 2))
            cell.number_format = self.QTY_FORMAT
            cell.alignment = Alignment(horizontal="right")
            if row_fill:
                cell.fill = row_fill
            cell.border = self.THIN_BORDER

            # Currency
            cell = ws.cell(row=row_idx, column=5, value=entry.currency or "")
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(color="555555", size=9)
            if row_fill:
                cell.fill = row_fill
            cell.border = self.THIN_BORDER

            # Avg. Unit Rate
            cell = ws.cell(row=row_idx, column=6, value=round(entry.unit_rate, 2))
            cell.number_format = self.COST_FORMAT
            cell.alignment = Alignment(horizontal="right")
            if row_fill:
                cell.fill = row_fill
            cell.border = self.THIN_BORDER

            # Total Cost
            cell = ws.cell(row=row_idx, column=7, value=round(entry.total_cost, 2))
            cell.number_format = self.COST_FORMAT
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True)
            if row_fill:
                cell.fill = row_fill
            cell.border = self.THIN_BORDER

            # Used In
            codes = ", ".join(sorted(entry.used_in_codes))
            cell = ws.cell(row=row_idx, column=8, value=codes)
            cell.font = Font(color="777777", size=9)
            if row_fill:
                cell.fill = row_fill
            cell.border = self.THIN_BORDER

        # Totals row
        if entries:
            total_row = len(entries) + 2
            total_qty = sum(e.total_qty for e in entries)
            total_cost = sum(e.total_cost for e in entries)

            ws.cell(row=total_row, column=1, value="").fill = self.TOTAL_FILL
            cell = ws.cell(row=total_row, column=2, value="TOTAL")
            cell.font = self.TOTAL_FONT
            cell.fill = self.TOTAL_FILL
            ws.cell(row=total_row, column=3, value="").fill = self.TOTAL_FILL

            cell = ws.cell(row=total_row, column=4, value=round(total_qty, 2))
            cell.number_format = self.QTY_FORMAT
            cell.alignment = Alignment(horizontal="right")
            cell.font = self.TOTAL_FONT
            cell.fill = self.TOTAL_FILL

            ws.cell(row=total_row, column=5, value="").fill = self.TOTAL_FILL
            ws.cell(row=total_row, column=6, value="").fill = self.TOTAL_FILL

            cell = ws.cell(row=total_row, column=7, value=round(total_cost, 2))
            cell.number_format = self.COST_FORMAT
            cell.alignment = Alignment(horizontal="right")
            cell.font = self.TOTAL_FONT
            cell.fill = self.TOTAL_FILL

            ws.cell(row=total_row, column=8, value="").fill = self.TOTAL_FILL

    # ── Summary Sheet ─────────────────────────────────────────────────────

    def _write_summary_sheet(self, wb):
        """Writes a summary worksheet with category totals."""
        ws = wb.create_sheet(title="Summary")
        s = self.result.summary

        # Title
        cell = ws.cell(row=1, column=1, value="Resources Schedule — Summary")
        cell.font = Font(name="Arial", bold=True, size=14, color="2E7D32")
        ws.merge_cells("A1:D1")

        # Headers
        headers = ["Resource Category", "Count", "Total Cost", "% of Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.SUMMARY_HEADER_FILL
            cell.font = self.SUMMARY_HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 14

        # Data rows
        grand = s['grand_total'] or 1  # avoid div/0
        categories = [
            ("Materials", s['materials_count'], s['materials_total'], self.MATERIAL_FILL),
            ("Labour", s['labor_count'], s['labor_total'], self.LABOR_FILL),
            ("Equipment", s['equipment_count'], s['equipment_total'], self.EQUIPMENT_FILL),
            ("Plant", s['plant_count'], s['plant_total'], self.PLANT_FILL),
        ]

        for i, (name, count, cost, fill) in enumerate(categories, 4):
            cell = ws.cell(row=i, column=1, value=name)
            cell.fill = fill
            cell.font = Font(bold=True)
            cell.border = self.THIN_BORDER

            cell = ws.cell(row=i, column=2, value=count)
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

            cell = ws.cell(row=i, column=3, value=round(cost, 2))
            cell.number_format = self.COST_FORMAT
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True)
            cell.border = self.THIN_BORDER

            pct = (cost / grand) * 100 if grand > 0 else 0
            cell = ws.cell(row=i, column=4, value=f"{pct:.1f}%")
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        # Grand total row
        total_row = 8
        total_count = sum(c[1] for c in categories)
        cell = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
        cell.font = Font(name="Arial", bold=True, size=11)
        cell.fill = self.TOTAL_FILL
        cell.border = self.THIN_BORDER

        cell = ws.cell(row=total_row, column=2, value=total_count)
        cell.font = self.TOTAL_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.fill = self.TOTAL_FILL
        cell.border = self.THIN_BORDER

        cell = ws.cell(row=total_row, column=3, value=round(s['grand_total'], 2))
        cell.number_format = self.COST_FORMAT
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(name="Arial", bold=True, size=11)
        cell.fill = self.TOTAL_FILL
        cell.border = self.THIN_BORDER

        cell = ws.cell(row=total_row, column=4, value="100.0%")
        cell.alignment = Alignment(horizontal="center")
        cell.font = self.TOTAL_FONT
        cell.fill = self.TOTAL_FILL
        cell.border = self.THIN_BORDER

        # Skipped items note
        if self.result.skipped_rows:
            note_row = total_row + 2
            cell = ws.cell(row=note_row, column=1,
                           value=f"⚠ {s['skipped_count']} items skipped (no resource breakdown)")
            cell.font = Font(color="C62828", italic=True)

        ws.freeze_panes = "A4"

    # ── Skipped Items Sheet ───────────────────────────────────────────────

    def _write_skipped_sheet(self, wb):
        """Writes a sheet listing all items that were skipped."""
        ws = wb.create_sheet(title="Skipped Items")

        headers = ["#", "Sheet", "Description", "Reason", "Bill Amount"]
        col_widths = [6, 15, 50, 30, 18]

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

        for row_idx, item in enumerate(self.result.skipped_rows, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=item.sheet)
            ws.cell(row=row_idx, column=3, value=item.description)

            cell = ws.cell(row=row_idx, column=4, value=item.reason)
            cell.font = Font(color="C62828")

            cell = ws.cell(row=row_idx, column=5, value=round(item.bill_amount, 2))
            cell.number_format = self.COST_FORMAT
            cell.alignment = Alignment(horizontal="right")
