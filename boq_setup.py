import os
import pandas as pd
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QSplitter, QTableWidget, QTableWidgetItem,
    QLabel, QComboBox, QPushButton, QTreeWidget, QTreeWidgetItem, QHeaderView,
    QMessageBox, QInputDialog, QMenu, QFormLayout, QTabWidget, QDockWidget,
    QScrollArea, QGroupBox, QListWidget, QAbstractItemView, QSizePolicy, QCheckBox,
    QListWidgetItem
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont

import pboq_constants as const

class BOQToolsPane(QWidget):
    """Encapsulates the tools for BOQ Setup into a scrollable pane."""
    def __init__(self, owner):
        super().__init__()
        self.owner = owner
        self.setMinimumWidth(250)
        self.setMaximumWidth(280)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        
        container = QWidget()
        container.setStyleSheet("font-size: 8pt;") # Slightly smaller font for compact look
        c_layout = QVBoxLayout(container)
        c_layout.setContentsMargins(5, 5, 5, 5)
        c_layout.setSpacing(5)

        # 1. Column Mapping
        col_group = QGroupBox("Column Mapping")
        col_layout = QFormLayout(col_group)
        col_layout.setContentsMargins(5, 5, 5, 5)
        col_layout.setSpacing(2)
        
        for cb in [self.owner.cb_ref, self.owner.cb_desc, self.owner.cb_qty, self.owner.cb_unit, self.owner.cb_rate]:
            cb.setMaximumHeight(20)
            
        col_layout.addRow("Ref / Item No:", self.owner.cb_ref)
        col_layout.addRow("Description:", self.owner.cb_desc)
        col_layout.addRow("Quantity:", self.owner.cb_qty)
        col_layout.addRow("Unit:", self.owner.cb_unit)
        col_layout.addRow("Rate (Optional):", self.owner.cb_rate)
        c_layout.addWidget(col_group)

        # 2. Sheet Selection
        sheet_group = QGroupBox("Sheets to Process")
        sheet_layout = QVBoxLayout(sheet_group)
        sheet_layout.setContentsMargins(5, 5, 5, 5)
        
        self.owner.sheet_selector.setMinimumHeight(120)
        self.owner.sheet_selector.setFixedWidth(240)
        sheet_layout.addWidget(self.owner.sheet_selector)
        c_layout.addWidget(sheet_group)

        # 3. Level Filter
        level_group = QGroupBox("Level Filter")
        level_layout = QVBoxLayout(level_group)
        level_layout.setContentsMargins(5, 5, 5, 5)
        level_layout.setSpacing(2)
        
        for i in range(1, 6):
            cb = self.owner.level_checkboxes[i]
            level_layout.addWidget(cb)
        c_layout.addWidget(level_group)

        # 4. Actions
        action_group = QGroupBox("Actions")
        action_layout = QVBoxLayout(action_group)
        action_layout.setContentsMargins(5, 12, 5, 5) # Increased top margin to prevent title overlap
        action_layout.setSpacing(4)
        
        # Apply Map Button (Greenish)
        self.owner.apply_map_btn.setMinimumHeight(30)
        self.owner.apply_map_btn.setStyleSheet("background-color: #2E7D32; color: white; font-weight: bold;")
        action_layout.addWidget(self.owner.apply_map_btn)
        
        self.owner.concat_btn.setMinimumHeight(30)
        action_layout.addWidget(self.owner.concat_btn)
        
        save_state_btn = QPushButton("Save State")
        save_state_btn.setMinimumHeight(30)
        save_state_btn.clicked.connect(self.owner._save_state)
        action_layout.addWidget(save_state_btn)
        
        save_sor_btn = QPushButton("Save to SOR")
        save_sor_btn.setMinimumHeight(30)
        save_sor_btn.clicked.connect(self.owner._save_to_sor)
        action_layout.addWidget(save_sor_btn)
        
        # Save PBOQ (Blue)
        self.owner.save_pboq_btn.setMinimumHeight(35)
        self.owner.save_pboq_btn.setStyleSheet("background-color: #1976D2; color: white; font-weight: bold;")
        action_layout.addWidget(self.owner.save_pboq_btn)
        
        c_layout.addWidget(action_group)
        c_layout.addStretch()
        
        scroll.setWidget(container)
        layout.addWidget(scroll)

class BOQSetupWindow(QWidget):
    """Parses Excel BOQs, allows mapping, and formats them into an Estimate struct."""
    def __init__(self, boq_file_path, active_est_window=None, parent=None, project_dir=None):
        super().__init__(parent)
        self.main_window = parent
        self.boq_file_path = boq_file_path
        self.active_est_window = active_est_window
        self.project_dir = project_dir
        
        # Dictionary to store dataframe and row types per sheet name
        self.sheet_data = {} 
        self.active_sheet = None
        self.concat_descriptions = False
        
        # Color codes for visual feedback
        self.COLOR_HEADING = const.COLOR_HEADING 
        self.COLOR_ITEM = const.COLOR_ITEM    
        self.COLOR_IGNORE = QColor("#ffebee")  # Light red

        self.setWindowTitle(f"BOQ Setup - {os.path.basename(boq_file_path)}")
        self.setMinimumSize(950, 400)
        self._init_ui()
        self._load_excel()

    def _init_ui(self):
        # Create all tools first so they can be referenced
        self.cb_ref = QComboBox()
        self.cb_desc = QComboBox()
        self.cb_qty = QComboBox()
        self.cb_unit = QComboBox()
        self.cb_rate = QComboBox()
        
        self.sheet_selector = QListWidget()
        self.sheet_selector.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.sheet_selector.itemSelectionChanged.connect(self._build_tree_preview)
        
        self.level_checkboxes = {}
        for i in range(1, 6):
            cb = QCheckBox(f"Level {i}")
            cb.setChecked(True if i <= 2 else False)
            cb.stateChanged.connect(self._build_tree_preview)
            self.level_checkboxes[i] = cb
            
        self.apply_map_btn = QPushButton("Apply Mapping")
        self.apply_map_btn.clicked.connect(self._apply_mapping)
        
        self.concat_btn = QPushButton("Concatenate Descriptions")
        self.concat_btn.clicked.connect(self._toggle_concatenate)
        
        self.save_pboq_btn = QPushButton("Save to Priced BOQ")
        self.save_pboq_btn.clicked.connect(self._save_to_priced_boq)
        
        # Main Layout
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(10, 5, 10, 10)
        
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(4)
        splitter.setStyleSheet("QSplitter::handle { background-color: #cccccc; border-radius: 2px; }")
        
        # LEFT PANE: Raw Excel View with Tabs
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        
        # File selector setup
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("Select BOQ File:"))
        self.boq_file_selector = QComboBox()
        self.boq_file_selector.setMinimumWidth(250)
        
        project_folder = getattr(self, 'project_dir', None)
        if not project_folder:
            project_folder = os.path.dirname(self.active_est_window.db_path) if self.active_est_window and getattr(self.active_est_window, 'db_path', None) else os.path.dirname(self.boq_file_path)
            if project_folder and os.path.basename(project_folder) == "Project Database":
                project_folder = os.path.dirname(project_folder)
                
        boq_dir = os.path.join(project_folder, "Imported BOQs")
        if os.path.exists(boq_dir):
            for f in os.listdir(boq_dir):
                if f.lower().endswith(('.xlsx', '.xls')):
                    self.boq_file_selector.addItem(f, os.path.join(boq_dir, f))
                    
        idx = self.boq_file_selector.findData(self.boq_file_path)
        if idx >= 0: self.boq_file_selector.setCurrentIndex(idx)
        else:
            self.boq_file_selector.addItem(os.path.basename(self.boq_file_path), self.boq_file_path)
            self.boq_file_selector.setCurrentIndex(self.boq_file_selector.count() - 1)
            
        self.boq_file_selector.activated.connect(self._on_boq_file_changed)
        file_layout.addWidget(self.boq_file_selector, stretch=1)
        left_layout.addLayout(file_layout)
        
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.TabPosition.South)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        left_layout.addWidget(self.tabs)
        
        # RIGHT PANE: Formatted Preview (Expanded)
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Sheet", "Ref", "Description", "Quantity", "Unit", "Level", "Type"])
        self.tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tree.header().setStretchLastSection(True)
        right_layout.addWidget(self.tree)
        
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(0, 5)
        splitter.setStretchFactor(1, 5)
        
        self.main_layout.addWidget(splitter)

        # Setup Tools Pane in Dock
        self.tools_pane = BOQToolsPane(self)
        if self.main_window:
            self.tools_dock = QDockWidget("BOQ Setup Tools", self.main_window)
            self.tools_dock.setWidget(self.tools_pane)
            self.main_window.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea, self.tools_dock)
            self.main_window.mdi_area.subWindowActivated.connect(self._on_mdi_subwindow_activated)
            self.destroyed.connect(self._cleanup_tools_dock)

    def _on_boq_file_changed(self, index):
        if index < 0: return
        
        new_path = self.boq_file_selector.itemData(index)
        if new_path and new_path != self.boq_file_path:
            self.boq_file_path = new_path
            self.setWindowTitle(f"BOQ Setup - {os.path.basename(new_path)}")
            self._load_excel()

    def _has_saved_state(self):
        """Check if a saved state file exists for the current BOQ file."""
        import json, os
        project_dir = getattr(self, 'project_dir', None)
        if not project_dir:
            project_dir = os.path.dirname(self.active_est_window.db_path) if self.active_est_window and getattr(self.active_est_window, 'db_path', None) else os.path.dirname(self.boq_file_path)
            if project_dir and os.path.basename(project_dir) == "Project Database":
                project_dir = os.path.dirname(project_dir)
        states_folder = os.path.join(project_dir, "BOQ-Setup States")
        base_name = os.path.basename(self.boq_file_path)
        state_file = os.path.join(states_folder, base_name + ".state.json")
        return os.path.exists(state_file)

    def _load_excel(self):
        try:
            from PyQt6.QtWidgets import QProgressDialog, QApplication
            from PyQt6.QtCore import Qt
            
            # Check if we have a saved state BEFORE doing heavy work
            has_state = self._has_saved_state()
            
            progress = QProgressDialog(
                "Restoring saved BOQ state..." if has_state else "Opening Excel file (this may take a moment)...",
                "Cancel", 0, 100, self
            )
            progress.setWindowTitle("Loading BOQ")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0) # Show immediately
            progress.setValue(5)
            QApplication.processEvents()
            
            # Clear existing data structures for a fresh load
            self.tabs.clear()
            self.sheet_selector.clear()
            self.sheet_data = {}
            self.active_sheet = None
            self.tree.clear()
            
            if progress.wasCanceled(): return
            
            xl = pd.ExcelFile(self.boq_file_path)
            sheet_names = xl.sheet_names
            
            progress.setValue(15)
            QApplication.processEvents()
            
            # Only load openpyxl/xlrd for bold detection if NO saved state exists.
            # When a state exists, the saved row_types are the definitive source of truth,
            # so the expensive formatting scan is unnecessary.
            is_xlsx = self.boq_file_path.lower().endswith('.xlsx')
            wb = None
            if not has_state:
                if is_xlsx:
                    import openpyxl
                    try:
                        wb = openpyxl.load_workbook(self.boq_file_path, data_only=True)
                    except: pass
                else:
                    import xlrd
                    try:
                        wb = xlrd.open_workbook(self.boq_file_path, formatting_info=True)
                    except: pass
                
            total_sheets = max(1, len(sheet_names))

            for s_idx, sheet_name in enumerate(sheet_names):
                if progress.wasCanceled(): break
                
                base_progress = 15 + (s_idx / total_sheets) * 85
                progress.setLabelText(f"Loading sheet {s_idx + 1} of {total_sheets}: {sheet_name}...")
                progress.setValue(int(base_progress))
                QApplication.processEvents()
                
                df = xl.parse(sheet_name, header=None)
                # Cast to object type before filling to prevent pandas ValueError when replacing float NaNs with strings
                df = df.astype(object)
                df.fillna("", inplace=True)
                row_types = ['ignore'] * len(df)
                
                table = QTableWidget()
                table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
                table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
                table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
                table.customContextMenuRequested.connect(self._show_context_menu)
                # Word wrap to handle long descriptions
                table.setWordWrap(True)
                
                # Suspend visual updates while populating the table for a massive speed boost
                table.setUpdatesEnabled(False)
                
                table.setRowCount(len(df))
                table.setColumnCount(len(df.columns))
                
                columns = [f"Column {i}" for i in range(len(df.columns))]
                table.setHorizontalHeaderLabels(columns)
                
                # Apply heading color and blue text to the header items directly
                # This is more robust than Palette when a global stylesheet is present
                for i in range(table.columnCount()):
                    item = table.horizontalHeaderItem(i)
                    if item:
                        item.setBackground(const.COLOR_HEADING)
                        item.setForeground(const.COLOR_HEADER_TEXT)
                        font = item.font()
                        font.setBold(True)
                        item.setFont(font)
                
                ws = None
                if wb and is_xlsx and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                elif wb and not is_xlsx:
                    try:
                        ws = wb.sheet_by_name(sheet_name)
                    except: pass
                    
                bold_font = QFont()
                bold_font.setBold(True)
                
                num_rows = len(df)
                num_cols = len(df.columns)
                
                for r in range(num_rows):
                    # Update progress every 100 rows
                    if r % 100 == 0:
                        if progress.wasCanceled(): break
                        current_progress = base_progress + ((r / num_rows) * (85 / total_sheets))
                        progress.setValue(int(current_progress))
                        QApplication.processEvents()
                        
                    row_is_bold = False
                    for c in range(num_cols):
                        val = str(df.iloc[r, c])
                        item = QTableWidgetItem(val)
                        
                        # Only do the expensive bold check if no saved state (first-time load)
                        if ws:
                            is_bold = False
                            if is_xlsx:
                                try:
                                    cell = ws.cell(row=r+1, column=c+1)
                                    if cell.font and cell.font.bold: is_bold = True
                                except: pass
                            else:
                                try:
                                    xf_idx = ws.cell_xf_index(r, c)
                                    xf = wb.xf_list[xf_idx]
                                    font = wb.font_list[xf.font_index]
                                    if font.weight >= 700: is_bold = True
                                except: pass
                                
                            if is_bold:
                                item.setFont(bold_font)
                                if val.strip():
                                    row_is_bold = True
                                
                        item.setBackground(self.COLOR_IGNORE)
                        table.setItem(r, c, item)
                        
                    # Auto guess: if major cell in row was bold, default it to heading initially
                    if row_is_bold:
                        row_types[r] = 'heading'
                        for c in range(num_cols):
                            if table.item(r, c): 
                                table.item(r, c).setBackground(self.COLOR_HEADING)
                
                # Re-enable updates before resize so Qt can compute layout once
                table.setUpdatesEnabled(True)
                
                # Automatically align column widths to content
                table.resizeColumnsToContents()
                # Enforce a maximum column width so that long descriptions are forced to wrap
                for c in range(table.columnCount()):
                    if table.columnWidth(c) > 400:
                        table.setColumnWidth(c, 400)
                
                # Resize rows to fit wrapped content based on new capped column widths
                table.resizeRowsToContents()
                # Automatically resize row height as user resizes columns/modifies layout
                table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
                
                self.sheet_data[sheet_name] = {
                    'df': df,
                    'row_types': row_types,
                    'table': table,
                    'columns': columns
                }
                self.tabs.addTab(table, sheet_name)
                
                # Add to selector list
                from PyQt6.QtWidgets import QListWidgetItem
                item = QListWidgetItem(sheet_name)
                self.sheet_selector.addItem(item)
                item.setSelected(True) # Select all by default

            if sheet_names and not progress.wasCanceled():
                self.active_sheet = sheet_names[0]
                self._populate_comboboxes(self.sheet_data[self.active_sheet]['columns'])
                
                self._load_saved_state()
                
            progress.setValue(100)

        except Exception as e:
            QMessageBox.critical(self, "Excel Error", f"Failed to load BOQ Excel file.\nError: {e}")

    def _on_tab_changed(self, index):
        sheet_name = self.tabs.tabText(index)
        self.active_sheet = sheet_name
        
    def _populate_comboboxes(self, columns):
        # Determine the maximum number of columns across all sheets to provide a stable list (e.g., 0-10 or 0-max)
        max_cols = max([len(data['columns']) for data in self.sheet_data.values()]) if self.sheet_data else 20
        
        explicit_columns = [f"Column {i}" for i in range(max_cols)]
        
        for cb in [self.cb_ref, self.cb_desc, self.cb_qty, self.cb_unit, self.cb_rate]:
            cb.clear()
            cb.addItem("-- Select Column --")
            cb.addItems(explicit_columns)
        
        # Set some reasonable defaults based on common BOQ layouts if possible
        if max_cols > 0: self.cb_ref.setCurrentIndex(1)
        if max_cols > 1: self.cb_desc.setCurrentIndex(2)
        if max_cols > 2: self.cb_qty.setCurrentIndex(3)
        if max_cols > 3: self.cb_unit.setCurrentIndex(4)

    def _show_context_menu(self, pos):
        if not self.active_sheet: return
        data = self.sheet_data[self.active_sheet]
        table = data['table']
        
        menu = QMenu()
        set_heading_action = menu.addAction("Set row(s) as Heading")
        set_item_action = menu.addAction("Set row(s) as Item")
        set_ignore_action = menu.addAction("Set row(s) to Ignore")
        
        action = menu.exec(table.mapToGlobal(pos))
        if not action: return
        
        rows = list(set([idx.row() for idx in table.selectedIndexes()]))
        
        for r in rows:
            if action == set_heading_action:
                data['row_types'][r] = 'heading'
                self._update_row_color(table, r, self.COLOR_HEADING)
            elif action == set_item_action:
                data['row_types'][r] = 'item'
                self._update_row_color(table, r, self.COLOR_ITEM)
            elif action == set_ignore_action:
                data['row_types'][r] = 'ignore'
                self._update_row_color(table, r, self.COLOR_IGNORE)
        
        self._build_tree_preview()

    def _update_row_color(self, table, row, color):
        for c in range(table.columnCount()):
            item = table.item(row, c)
            if item:
                item.setBackground(color)

    def _apply_mapping(self):
        """Auto detects Headings vs Items based on columns for selected sheets only."""
        desc_col = self.cb_desc.currentIndex() - 1
        qty_col = self.cb_qty.currentIndex() - 1
        
        if desc_col < 0:
            QMessageBox.warning(self, "Mapping Error", "You must at least select a Description column.")
            return

        selected_sheets = [item.text() for item in self.sheet_selector.selectedItems()]
        if not selected_sheets:
            QMessageBox.warning(self, "Warning", "Please select at least one sheet to apply mapping to.")
            return

        for sheet_name in selected_sheets:
            if sheet_name not in self.sheet_data: continue
            data = self.sheet_data[sheet_name]
            df = data['df']
            table = data['table']
            
            for r in range(len(df)):
                desc_val = str(df.iloc[r, desc_col]).strip() if 0 <= desc_col < len(df.columns) else ""
                qty_val = str(df.iloc[r, qty_col]).strip() if 0 <= qty_col < len(df.columns) else ""
                
                if desc_val.lower() == 'nan' or desc_val.lower() == '<na>': desc_val = ""
                if qty_val.lower() == 'nan' or qty_val.lower() == '<na>': qty_val = ""
                
                # We skip overriding if the user manually set it, but for simplicity we'll override if ignore
                # Actually, let's keep existing headings derived from bold font if they exist.
                if not desc_val:
                    data['row_types'][r] = 'ignore'
                    self._update_row_color(table, r, self.COLOR_IGNORE)
                    continue
                    
                # Check for table header row
                is_table_header = False
                desc_lower = desc_val.lower()
                qty_lower = qty_val.lower()
                if desc_lower in ['description', 'item description', 'item', 'description of work'] or \
                   qty_lower in ['qty', 'quantity', 'unit', 'rate', 'amount', 'bill amount']:
                    is_table_header = True
                    
                # If it has a description but no quantity, or it's a table header, assume Heading
                if is_table_header or (desc_val and not qty_val):
                    data['row_types'][r] = 'heading'
                    self._update_row_color(table, r, self.COLOR_HEADING)
                elif desc_val and qty_val:
                    data['row_types'][r] = 'item'
                    self._update_row_color(table, r, self.COLOR_ITEM)

        self._build_tree_preview()

    def _toggle_concatenate(self):
        self.concat_descriptions = not self.concat_descriptions
        if self.concat_descriptions:
            self.concat_btn.setText("Un-Catenate Descriptions")
        else:
            self.concat_btn.setText("Concatenate Descriptions")
        self._build_tree_preview()

    def _build_tree_preview(self):
        self.tree.clear()
        
        is_concat = getattr(self, 'concat_descriptions', False)
        # Toggle visibility of Level and Type columns based on concatenation state
        self.tree.setColumnHidden(5, is_concat)
        self.tree.setColumnHidden(6, is_concat)
        
        ref_col = self.cb_ref.currentIndex() - 1
        desc_col = self.cb_desc.currentIndex() - 1
        qty_col = self.cb_qty.currentIndex() - 1
        unit_col = self.cb_unit.currentIndex() - 1
        
        bold_font = QFont()
        bold_font.setBold(True)
        
        selected_sheets = [item.text() for item in self.sheet_selector.selectedItems()]
        
        selected_levels = [lvl for lvl, cb in self.level_checkboxes.items() if cb.isChecked()]

        for sheet_name in selected_sheets:
            if sheet_name not in self.sheet_data: continue
            data = self.sheet_data[sheet_name]
            df = data['df']
            row_types = data['row_types']
            
            # Dynamically calculate levels based on current row_types
            levels = [0] * len(df)
            current_heading_level = 0
            for r in range(len(df) - 1, -1, -1):
                rtype = row_types[r]
                if rtype == 'item':
                    levels[r] = 1
                    current_heading_level = 2
                elif rtype == 'heading':
                    if current_heading_level > 0:
                        levels[r] = current_heading_level
                        current_heading_level += 1
                    else:
                        levels[r] = 0
            data['levels'] = levels
            
            # Create a root node for the sheet
            sheet_node = QTreeWidgetItem(self.tree, [sheet_name, "", "Sheet Root", "", "", "", "Heading"])
            for i in range(7): sheet_node.setFont(i, bold_font)
            sheet_node.setBackground(2, QColor("#bbdefb")) # light blue

            current_heading_item = sheet_node
            current_headings = {} # Map to track active headings keyed by their level
            
            for r in range(len(df)):
                rtype = row_types[r]
                level = levels[r]
                if rtype == 'ignore' or level not in selected_levels: 
                    continue
                
                ref_val = str(df.iloc[r, ref_col]).strip() if 0 <= ref_col < len(df.columns) else ""
                desc_val = str(df.iloc[r, desc_col]).strip() if 0 <= desc_col < len(df.columns) else ""
                qty_val = str(df.iloc[r, qty_col]).strip() if 0 <= qty_col < len(df.columns) else ""
                unit_val = str(df.iloc[r, unit_col]).strip() if 0 <= unit_col < len(df.columns) else ""
                
                if ref_val.lower() in ('nan', '<na>'): ref_val = ""
                if desc_val.lower() in ('nan', '<na>'): desc_val = ""
                if qty_val.lower() in ('nan', '<na>'): qty_val = ""
                if unit_val.lower() in ('nan', '<na>'): unit_val = ""
                
                level_str = str(level) if level > 0 else ""

                if rtype == 'heading':
                    # Update the active headings dictionary and flush any lower-level headings
                    current_headings[level] = desc_val
                    keys_to_delete = [k for k in current_headings.keys() if k < level]
                    for k in keys_to_delete:
                        del current_headings[k]
                    
                    if not is_concat:
                        current_heading_item = QTreeWidgetItem(sheet_node, [sheet_name, ref_val, desc_val, "", "", level_str, "Heading"])
                        for i in range(7): current_heading_item.setFont(i, bold_font)
                        current_heading_item.setBackground(2, self.COLOR_HEADING)
                    else:
                        current_heading_item = sheet_node
                
                elif rtype == 'item':
                    # Prepare concatenated description dynamically capturing all active filtered headings
                    item_desc = desc_val
                    if is_concat:
                        active_heading_texts = [current_headings[lvl] for lvl in sorted(current_headings.keys(), reverse=True) if lvl in selected_levels]
                        if active_heading_texts:
                            item_desc = " : ".join(active_heading_texts) + f" : {desc_val}"
                        
                    # Flatten hierarchy if concatenated
                    parent = sheet_node if is_concat else (current_heading_item if current_heading_item else sheet_node)
                    
                    item_node = QTreeWidgetItem(parent, [sheet_name, ref_val, item_desc, qty_val, unit_val, level_str, "Item"])
                    item_node.setBackground(2, self.COLOR_ITEM)
                    
        self.tree.expandAll()

    def _save_to_priced_boq(self):
        """Creates a SQLite DB for the raw BOQ data (from Excel) in a 'Priced BOQs' folder."""
        import os
        import sqlite3
        import pandas as pd
        import json
        
        if not self.sheet_data:
            QMessageBox.warning(self, "Warning", "No data loaded to save.")
            return
        
        project_folder = getattr(self, 'project_dir', None)
        if not project_folder:
            project_folder = os.path.dirname(self.active_est_window.db_path) if self.active_est_window and getattr(self.active_est_window, 'db_path', None) else os.path.dirname(self.boq_file_path)
            if project_folder and os.path.basename(project_folder) == "Project Database":
                project_folder = os.path.dirname(project_folder)
            
        pboq_folder = os.path.join(project_folder, "Priced BOQs")
        if not os.path.exists(pboq_folder):
            try:
                os.makedirs(pboq_folder)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to create Priced BOQs folder:\n{e}")
                return
                
        base_name = os.path.basename(self.boq_file_path)
        name, _ = os.path.splitext(base_name)
        new_file_name = f"PBOQ_{name}.db"
        pboq_file_path = os.path.join(pboq_folder, new_file_name)
        
        # Load Excel workbook for formatting extraction
        is_xlsx = self.boq_file_path.lower().endswith('.xlsx')
        wb = None
        if is_xlsx:
            import openpyxl
            try:
                wb = openpyxl.load_workbook(self.boq_file_path, data_only=True)
            except: pass
        else:
            import xlrd
            try:
                wb = xlrd.open_workbook(self.boq_file_path, formatting_info=True)
            except: pass
        
        # Collect raw data and formatting from all sheets
        all_rows = []
        all_formats = []
        max_cols = 0
        global_row_idx = 0
        
        for sheet_name, data in self.sheet_data.items():
            df = data['df']
            max_cols = max(max_cols, len(df.columns))
            
            # Get the worksheet for formatting
            ws = None
            if wb and is_xlsx and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb and not is_xlsx:
                try:
                    ws = wb.sheet_by_name(sheet_name)
                except: pass
            
            for r in range(len(df)):
                row_values = []
                for c in range(len(df.columns)):
                    val = str(df.iloc[r, c])
                    if val.lower() in ('nan', '<na>', 'none'):
                        val = ""
                    row_values.append(val)
                    
                    # Extract formatting
                    fmt = {}
                    if ws and is_xlsx:
                        try:
                            cell = ws.cell(row=r+1, column=c+1)
                            if cell.font:
                                if cell.font.bold: fmt['bold'] = True
                                if cell.font.italic: fmt['italic'] = True
                                if cell.font.underline: fmt['underline'] = True
                                if cell.font.color and cell.font.color.rgb and cell.font.color.rgb != '00000000':
                                    rgb = str(cell.font.color.rgb)
                                    if len(rgb) == 8: rgb = rgb[2:]  # Strip alpha
                                    fmt['font_color'] = f"#{rgb}"
                            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                                rgb = str(cell.fill.fgColor.rgb)
                                if rgb and rgb != '00000000':
                                    if len(rgb) == 8: rgb = rgb[2:]
                                    fmt['bg_color'] = f"#{rgb}"
                        except: pass
                    elif ws and not is_xlsx:
                        try:
                            xf_idx = ws.cell_xf_index(r, c)
                            xf = wb.xf_list[xf_idx]
                            font = wb.font_list[xf.font_index]
                            if font.weight >= 700: fmt['bold'] = True
                            if font.italic: fmt['italic'] = True
                            if font.underlined: fmt['underline'] = True
                        except: pass
                    
                    if fmt:
                        all_formats.append({
                            'row_idx': global_row_idx,
                            'col_idx': c,
                            'fmt_json': json.dumps(fmt)
                        })
                
                all_rows.append({"Sheet": sheet_name, "row_values": row_values})
                global_row_idx += 1
        
        if not all_rows:
            QMessageBox.warning(self, "Warning", "No data to save to Priced BOQ.")
            return
        
        # Build records with generic column names: Sheet + Column 0, Column 1, ...
        records = []
        for row_entry in all_rows:
            record = {"Sheet": row_entry["Sheet"]}
            for c_idx, val in enumerate(row_entry["row_values"]):
                record[f"Column {c_idx}"] = val
            # Fill remaining columns if this sheet had fewer columns
            for c_idx in range(len(row_entry["row_values"]), max_cols):
                record[f"Column {c_idx}"] = ""
            records.append(record)
            
        try:
            df_out = pd.DataFrame(records)
            conn = sqlite3.connect(pboq_file_path)
            df_out.to_sql('pboq_items', conn, if_exists='replace', index=False)
            
            # Save formatting data
            cursor = conn.cursor()
            cursor.execute("DROP TABLE IF EXISTS pboq_formatting")
            cursor.execute("""
                CREATE TABLE pboq_formatting (
                    row_idx INTEGER,
                    col_idx INTEGER,
                    fmt_json TEXT
                )
            """)
            if all_formats:
                cursor.executemany(
                    "INSERT INTO pboq_formatting (row_idx, col_idx, fmt_json) VALUES (?, ?, ?)",
                    [(f['row_idx'], f['col_idx'], f['fmt_json']) for f in all_formats]
                )
            conn.commit()
            conn.close()
            QMessageBox.information(self, "Success", f"Successfully saved Priced BOQ to:\n{pboq_file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Priced BOQ Database file:\n{e}")

    def _save_state(self):
        import json, os
        project_dir = getattr(self, 'project_dir', None)
        if not project_dir:
            project_dir = os.path.dirname(self.active_est_window.db_path) if self.active_est_window and getattr(self.active_est_window, 'db_path', None) else os.path.dirname(self.boq_file_path)
            if project_dir and os.path.basename(project_dir) == "Project Database":
                project_dir = os.path.dirname(project_dir)
            
        states_folder = os.path.join(project_dir, "BOQ-Setup States")
        os.makedirs(states_folder, exist_ok=True)
        
        base_name = os.path.basename(self.boq_file_path)
        state_file = os.path.join(states_folder, base_name + ".state.json")
        
        state = {
            'cb_ref': self.cb_ref.currentIndex(),
            'cb_desc': self.cb_desc.currentIndex(),
            'cb_qty': self.cb_qty.currentIndex(),
            'cb_unit': self.cb_unit.currentIndex(),
            'cb_rate': self.cb_rate.currentIndex(),
            'selected_sheets': [item.text() for item in self.sheet_selector.selectedItems()],
            'level_checkboxes': {lvl: cb.isChecked() for lvl, cb in self.level_checkboxes.items()}
        }
        
        row_types_dict = {}
        for sheet, data in self.sheet_data.items():
            row_types_dict[sheet] = data['row_types']
            
        state['row_types'] = row_types_dict
        
        try:
            with open(state_file, 'w') as f:
                json.dump(state, f)
            QMessageBox.information(self, "Success", "State saved successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save state:\n{e}")

    def _save_to_sor(self):
        import os
        from PyQt6.QtWidgets import QTreeWidgetItemIterator
        
        project_folder = getattr(self, 'project_dir', None)
        if not project_folder:
            project_folder = os.path.dirname(self.active_est_window.db_path) if self.active_est_window and getattr(self.active_est_window, 'db_path', None) else os.path.dirname(self.boq_file_path)
            if project_folder and os.path.basename(project_folder) == "Project Database":
                project_folder = os.path.dirname(project_folder)
            
        sor_folder = os.path.join(project_folder, "SOR")
        if not os.path.exists(sor_folder):
            try:
                os.makedirs(sor_folder)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to create SOR folder:\n{e}")
                return
                
        base_name = os.path.basename(self.boq_file_path)
        name, ext = os.path.splitext(base_name)
        new_file_name = f"SOR_{name}.db"
        sor_file_path = os.path.join(sor_folder, new_file_name)
        
        data = []
        iterator = QTreeWidgetItemIterator(self.tree)
        while iterator.value():
            item = iterator.value()
            desc = item.text(2)
            item_type = item.text(6)
            
            # Skip the artificial Sheet Root nodes
            if item_type == "Heading" and desc == "Sheet Root":
                iterator += 1
                continue
                
            data.append({
                "Sheet": item.text(0),
                "Ref": item.text(1),
                "Description": desc,
                "Quantity": item.text(3),
                "Unit": item.text(4)
            })
            iterator += 1
            
        if not data:
            QMessageBox.warning(self, "Warning", "No data to save to SOR. Please apply mapping first.")
            return
            
        try:
            import sqlite3
            df = pd.DataFrame(data)
            conn = sqlite3.connect(sor_file_path)
            df.to_sql('sor_items', conn, if_exists='replace', index=False)
            conn.close()
            QMessageBox.information(self, "Success", f"Successfully saved Formatted Preview to:\n{sor_file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save SOR Database file:\n{e}")

    def _load_saved_state(self):
        import json, os
        project_dir = getattr(self, 'project_dir', None)
        if not project_dir:
            project_dir = os.path.dirname(self.active_est_window.db_path) if self.active_est_window and getattr(self.active_est_window, 'db_path', None) else os.path.dirname(self.boq_file_path)
            if project_dir and os.path.basename(project_dir) == "Project Database":
                project_dir = os.path.dirname(project_dir)
            
        states_folder = os.path.join(project_dir, "BOQ-Setup States")
        
        base_name = os.path.basename(self.boq_file_path)
        state_file = os.path.join(states_folder, base_name + ".state.json")
        if not os.path.exists(state_file):
            return False
            
        try:
            with open(state_file, 'r') as f:
                state = json.load(f)
                
            if 'cb_ref' in state: self.cb_ref.setCurrentIndex(state['cb_ref'])
            if 'cb_desc' in state: self.cb_desc.setCurrentIndex(state['cb_desc'])
            if 'cb_qty' in state: self.cb_qty.setCurrentIndex(state['cb_qty'])
            if 'cb_unit' in state: self.cb_unit.setCurrentIndex(state['cb_unit'])
            if 'cb_rate' in state: self.cb_rate.setCurrentIndex(state['cb_rate'])
            
            if 'level_checkboxes' in state:
                for lvl_str, is_checked in state['level_checkboxes'].items():
                    lvl = int(lvl_str)
                    if lvl in self.level_checkboxes:
                        self.level_checkboxes[lvl].setChecked(is_checked)
                        
            if 'selected_sheets' in state:
                for i in range(self.sheet_selector.count()):
                    item = self.sheet_selector.item(i)
                    if item.text() in state['selected_sheets']:
                        item.setSelected(True)
                    else:
                        item.setSelected(False)
                        
            if 'row_types' in state:
                for sheet, row_types in state['row_types'].items():
                    if sheet in self.sheet_data:
                        self.sheet_data[sheet]['row_types'] = row_types
                        table = self.sheet_data[sheet]['table']
                        # Suspend repaints during bulk color update
                        table.setUpdatesEnabled(False)
                        num_cols = table.columnCount()
                        for r, rtype in enumerate(row_types):
                            color = self.COLOR_IGNORE
                            if rtype == 'heading': color = self.COLOR_HEADING
                            elif rtype == 'item': color = self.COLOR_ITEM
                            
                            for c in range(num_cols):
                                item = table.item(r, c)
                                if item: item.setBackground(color)
                        table.setUpdatesEnabled(True)
                                
            self._build_tree_preview()
            return True
        except Exception as e:
            print(f"Error loading state: {e}")
            return False

    def _on_mdi_subwindow_activated(self, sub):
        """Toggle dock visibility based on whether THIS window is active."""
        if not hasattr(self, 'tools_dock') or not self.tools_dock:
            return
        
        # 'sub' is the QMdiSubWindow. Its widget should be 'self' if it's active.
        if sub and sub.widget() == self:
            self.tools_dock.show()
            self.tools_dock.raise_()
        else:
            self.tools_dock.hide()

    def closeEvent(self, event):
        """Ensure the dock is hidden when closing the window."""
        try:
            if hasattr(self, 'tools_dock') and self.tools_dock:
                self.tools_dock.hide()
        except RuntimeError:
            pass
        super().closeEvent(event)

    def _cleanup_tools_dock(self):
        """Cleanup dock widget when the viewer is destroyed."""
        if self.main_window:
            try:
                if hasattr(self, 'tools_dock') and self.tools_dock:
                    # Remove it from the main window's layout
                    self.main_window.removeDockWidget(self.tools_dock)
                    self.tools_dock.deleteLater()
                    self.tools_dock = None
            except RuntimeError:
                pass
