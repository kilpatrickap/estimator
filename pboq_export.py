"""
PBOQ Excel Exporter — Exports a Priced BOQ database to a styled Excel (.xlsx) workbook.

Each BOQ sheet in the database becomes a separate worksheet. The export includes
both physical columns (Ref, Desc, Qty, Unit, Bill Rate, Bill Amount) and logical
columns (Gross Rate, Rate Code, Plug Rate, Plug Code, Sub. Rate, Sub. Code,
Prov Sum, Prov Sum Code, PC Sum, PC Sum Code, Daywork, Daywork Code).
"""

import os
import json
import sqlite3
from pboq_logic import PBOQLogic

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class PBOQExcelExporter:
    """Exports a PBOQ .db file to a styled .xlsx workbook."""

    # ── Column specification ──────────────────────────────────────────────
    # Each entry: (header_label, source_type, source_key, hex_fill, is_numeric)
    #   source_type: 'physical' = from physical columns via mapping index
    #                'logical'  = from logical column name in DB
    COLUMN_SPEC = [
        ("Ref",            "physical", "ref",            "e3f2fd", False),
        ("Description",    "physical", "desc",           "e3f2fd", False),
        ("Qty",            "physical", "qty",            "e3f2fd", True),
        ("Unit",           "physical", "unit",           "e3f2fd", False),
        ("Bill Rate",      "physical", "bill_rate",      "fff9c4", True),
        ("Bill Amount",    "physical", "bill_amount",    "fff9c4", True),
        ("Gross Rate",     "logical",  "GrossRate",      "e8f5e9", True),
        ("Rate Code",      "logical",  "RateCode",       "e8f5e9", False),
        ("Plug Rate",      "logical",  "PlugRate",       "f3e5f5", True),
        ("Plug Code",      "logical",  "PlugCode",       "f3e5f5", False),
        ("Sub. Rate",      "logical",  "SubbeeRate",     "ffe0b2", True),
        ("Sub. Code",      "logical",  "SubbeeCode",     "ffe0b2", False),
        ("Prov Sum",       "logical",  "ProvSum",        "E0FFFF", True),
        ("Prov Sum Code",  "logical",  "ProvSumCode",    "E0FFFF", False),
        ("PC Sum",         "logical",  "PCSum",          "D4FF99", True),
        ("PC Sum Code",    "logical",  "PCSumCode",      "D4FF99", False),
        ("Daywork",        "logical",  "Daywork",        "D7CCC8", True),
        ("Daywork Code",   "logical",  "DayworkCode",    "D7CCC8", False),
    ]

    # Header row style
    HEADER_FILL = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid") if HAS_OPENPYXL else None
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10) if HAS_OPENPYXL else None
    FLAGGED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid") if HAS_OPENPYXL else None
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="CCCCCC")
    ) if HAS_OPENPYXL else None

    def __init__(self, db_path, project_dir):
        """
        Args:
            db_path:     Full path to the PBOQ .db file.
            project_dir: Project root directory (contains 'PBOQ States' folder).
        """
        self.db_path = db_path
        self.project_dir = project_dir
        self.db_filename = os.path.basename(db_path)

    # ── Public API ────────────────────────────────────────────────────────

    def export(self, output_path):
        """
        Exports the PBOQ database to an Excel workbook.

        Args:
            output_path: Destination .xlsx file path.

        Returns:
            (True, message) on success, (False, error_message) on failure.
        """
        if not HAS_OPENPYXL:
            return False, "openpyxl is not installed. Run: pip install openpyxl"

        if not os.path.exists(self.db_path):
            return False, f"PBOQ database not found: {self.db_path}"

        try:
            # 1. Load column mappings from state file
            mappings = self._load_mappings()

            # 2. Load data from DB
            sheet_groups, db_columns, logical_col_names = self._load_data()
            if not sheet_groups:
                return False, "No data found in the PBOQ database."

            # 3. Build workbook
            wb = Workbook()
            # Remove the default sheet created by openpyxl
            wb.remove(wb.active)

            for sheet_name, rows in sheet_groups.items():
                ws = wb.create_sheet(title=self._sanitize_sheet_name(sheet_name))
                self._write_sheet(ws, rows, mappings, db_columns, logical_col_names)

            # 4. Save
            wb.save(output_path)
            return True, f"Exported {len(sheet_groups)} sheet(s) to:\n{output_path}"

        except Exception as e:
            return False, f"Export failed: {e}"

    # ── Private helpers ───────────────────────────────────────────────────

    def _load_mappings(self):
        """Loads column mappings from the PBOQ state JSON file."""
        state_path = os.path.join(
            self.project_dir, "PBOQ States", self.db_filename + ".json"
        )
        if os.path.exists(state_path):
            try:
                with open(state_path, "r") as f:
                    data = json.load(f)
                    return data.get("mappings", {})
            except Exception:
                pass
        return {}

    def _load_data(self):
        """
        Reads all rows from pboq_items, grouped by Sheet.

        Returns:
            sheet_groups: dict  {sheet_name: [(physical_data_list, logical_data_dict, is_flagged), ...]}
            db_columns:   list  of physical column names (including 'Sheet' at index 0)
            logical_col_names: list  of logical column names queried
        """
        conn = sqlite3.connect(self.db_path)
        PBOQLogic.ensure_schema(conn)
        cursor = conn.cursor()

        # Get physical columns
        cursor.execute("PRAGMA table_info(pboq_items)")
        all_db_cols = [info[1] for info in cursor.fetchall()]

        # The db_columns list matches what PBOQLogic.ensure_schema returns
        # Physical columns are the first N columns, starting with 'Sheet'
        # We need the same ordering as the viewer uses
        physical_cols = all_db_cols  # All columns from PRAGMA

        # Logical columns we need to export
        logical_col_names = [
            "GrossRate", "RateCode",
            "PlugRate", "PlugCode",
            "SubbeeRate", "SubbeeCode",
            "ProvSum", "ProvSumCode",
            "PCSum", "PCSumCode",
            "Daywork", "DayworkCode",
            "IsFlagged",
        ]

        # Build query: select physical + logical columns
        # Some logical columns are ALSO physical columns, so we need to be careful
        # to avoid duplicates in the SELECT
        quoted_physical = [f'"{c}"' for c in physical_cols]
        extra_logical = []
        for lc in logical_col_names:
            if lc not in physical_cols:
                extra_logical.append(f'"{lc}"')
            # If it's already in physical_cols, we'll read it from there

        query_parts = ["rowid"] + quoted_physical
        if extra_logical:
            query_parts += extra_logical

        query = f"SELECT {', '.join(query_parts)} FROM pboq_items"
        cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()

        # Build column index lookup for the query result
        result_col_names = ["rowid"] + physical_cols + [
            lc for lc in logical_col_names if lc not in physical_cols
        ]
        col_index = {name: i for i, name in enumerate(result_col_names)}

        # Group by Sheet
        sheet_groups = {}
        sheet_col_idx = col_index.get("Sheet", 1)  # 'Sheet' should be the first physical col

        for row in rows:
            sheet_name = str(row[sheet_col_idx]) if row[sheet_col_idx] else "Sheet 1"

            # Extract physical data (skip 'Sheet' at index 0 of physical cols)
            # physical_cols[0] is 'Sheet', physical_cols[1:] are display columns
            physical_data = []
            for pc in physical_cols:
                idx = col_index.get(pc)
                if idx is not None:
                    physical_data.append(row[idx])
                else:
                    physical_data.append("")

            # Extract logical data as dict
            logical_data = {}
            for lc in logical_col_names:
                idx = col_index.get(lc)
                if idx is not None:
                    logical_data[lc] = row[idx]
                else:
                    logical_data[lc] = None

            # Check flagged status
            flagged_val = logical_data.get("IsFlagged")
            is_flagged = flagged_val in [1, "1", True, "True"]

            if sheet_name not in sheet_groups:
                sheet_groups[sheet_name] = []

            sheet_groups[sheet_name].append((physical_data, logical_data, is_flagged))

        return sheet_groups, physical_cols, logical_col_names

    def _write_sheet(self, ws, rows, mappings, db_columns, logical_col_names):
        """Writes a single worksheet with headers, data, and styling."""
        num_cols = len(self.COLUMN_SPEC)

        # ── Write header row ──────────────────────────────────────────────
        for col_idx, (header, _, _, _, _) in enumerate(self.COLUMN_SPEC, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze header row
        ws.freeze_panes = "A2"

        # ── Build physical column index map ───────────────────────────────
        # db_columns includes 'Sheet' at [0], display columns start at [1]
        # mappings values are 0-based indices into display columns (i.e., db_columns[1:])
        # So mapping value N corresponds to db_columns[N + 1]
        display_cols = db_columns[1:] if len(db_columns) > 1 else []

        # ── Write data rows ──────────────────────────────────────────────
        # Track max content widths for auto-fit
        col_widths = [len(spec[0]) + 2 for spec in self.COLUMN_SPEC]

        for row_offset, (physical_data, logical_data, is_flagged) in enumerate(rows):
            excel_row = row_offset + 2  # Row 1 is header

            for col_idx, (header, source_type, source_key, hex_fill, is_numeric) in enumerate(self.COLUMN_SPEC, start=1):
                raw_value = None

                if source_type == "physical":
                    # Get the display column index from mappings
                    mapped_idx = mappings.get(source_key, -1)
                    if mapped_idx >= 0 and mapped_idx < len(display_cols):
                        # physical_data[0] is 'Sheet', so display col N = physical_data[N + 1]
                        phys_idx = mapped_idx + 1
                        if phys_idx < len(physical_data):
                            raw_value = physical_data[phys_idx]
                else:
                    # Logical column — read from logical_data dict
                    raw_value = logical_data.get(source_key)

                # Convert to proper type
                cell_value = self._clean_value(raw_value, is_numeric)
                cell = ws.cell(row=excel_row, column=col_idx, value=cell_value)

                # ── Cell styling ──
                # Background fill
                if is_flagged and col_idx <= 2:
                    # Flagged rows get red tint on Ref/Desc columns
                    cell.fill = self.FLAGGED_FILL
                else:
                    cell.fill = PatternFill(
                        start_color=hex_fill, end_color=hex_fill, fill_type="solid"
                    )

                # Number format
                if is_numeric and isinstance(cell_value, (int, float)):
                    cell.number_format = '#,##0.00'

                # Alignment
                if is_numeric:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif source_key == "unit":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Font
                cell.font = Font(name="Arial", size=10)

                # Subtle row border
                cell.border = self.THIN_BORDER

                # Track column width
                display_len = len(str(cell_value)) if cell_value else 0
                if display_len > col_widths[col_idx - 1]:
                    col_widths[col_idx - 1] = min(display_len, 50)

        # ── Auto-fit column widths ────────────────────────────────────────
        for col_idx in range(1, num_cols + 1):
            # Add a small padding
            width = col_widths[col_idx - 1] + 3
            ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)

    def _clean_value(self, raw, is_numeric):
        """Converts a raw DB value to an appropriate Python type for Excel."""
        if raw is None or str(raw).strip() == "" or str(raw).strip() == "None":
            return ""

        if is_numeric:
            text = str(raw).replace(",", "").replace(" ", "").strip()
            # Strip common currency symbols
            for sym in ["₵", "$", "€", "£", "¥", "R"]:
                text = text.replace(sym, "")
            try:
                return float(text)
            except (ValueError, TypeError):
                return str(raw)  # Fall back to text if not parseable

        return str(raw)

    def _sanitize_sheet_name(self, name):
        """Ensures the sheet name is valid for Excel (max 31 chars, no special chars)."""
        # Excel sheet name rules: max 31 chars, no []:*?/\
        clean = str(name)
        for ch in ["[", "]", ":", "*", "?", "/", "\\"]:
            clean = clean.replace(ch, "_")
        return clean[:31]
